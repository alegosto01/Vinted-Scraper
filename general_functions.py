from openpyxl import load_workbook
import os
import random as rnd
import time
import requests
from selenium.common.exceptions import WebDriverException
import pandas as pd
import re

def empty_excel(path):
    wb = load_workbook(path)
    ws = wb['Sheet1']  # Change 'Sheet1' to your specific sheet name

    # Loop through all rows and columns to clear content
    for row in ws.iter_rows(min_row=1, max_row=ws.max_row, min_col=1, max_col=ws.max_column):
        for cell in row:
            cell.value = None

    # Save the workbook
    wb.save(path)


    
def remove_illegal_characters(value):
    """Removes illegal characters that Excel does not support."""
    ILLEGAL_CHARACTERS = [chr(i) for i in range(32) if i not in (9, 10, 13)] + [chr(127)]
    for char in ILLEGAL_CHARACTERS:
        value = value.replace(char, "")
    return value


def ensure_path_exists(path):
    if not os.path.exists(path):
        os.makedirs(path)  # Create directories if they don't exist


def get_last_non_empty_row_excel(file_path):
    wb = load_workbook(file_path)
    ws = wb['Sheet1']  # Change 'Sheet1' to your sheet name
    for row in range(ws.max_row, 0, -1):
        if any(cell.value is not None for cell in ws[row]):
            return row
    return 0  # If all rows are empty

def random_sleep(range_from, range_to):
    seconds = rnd.uniform(range_from, range_to)
    time.sleep(seconds)



def download_image(image_url, path):
    response = requests.get(image_url)
    if not os.path.exists(path):
        with open(path, 'wb') as file:
            file.write(response.content)
    else:
        print("non scarico immagine esiste già")

   
def split_data(entry):
    # Split the entry by comma to separate title from other details
    title, details = entry.split(',', 1)

    # Extract individual components from details

    #   old way to get the price but they changed the html 

    # price = details.split('prezzo:')[1].split('€')[0].strip() + '€'  # Extract price #you can have other value so you probalby have to change zl to value that is in your country
# Scarpe, brand: Nike, condizioni: Nuovo con cartellino, taglia: 45, €110.00, €116.20 include la Protezione acquisti

    # print(f"DETAILS: {details}")
    split = details.split("€")
    if len(split) == 3:
        price = split[2].split('include la Protezione acquisti')[0]
    else:
        price = split[1]

    if "brand:" in details:
        brand = details.split('brand:')[1].split(',')[0].strip()  # Extract brand
    else:
        brand = "No brand"
    if "taglia:" in details:
        size = details.split('taglia:')[1].split(",")[0].strip()  # Extract size
    else:
        size = 0
    return title.strip(), price, brand, size


MAX_RETRIES = 3 # Maximum number of retries

def load_page(driver, webpage, page):
    retries = 0
    while retries < MAX_RETRIES:
        try:
            driver.get(f"{webpage}&page={page+1}")
            break  # Break the loop if the page loads successfully
        except WebDriverException as e:
            print(f"Error loading page: {e}")
            retries += 1
            time.sleep(2)  # Wait for 2 seconds before retrying
            if retries == MAX_RETRIES:
                print("Max retries reached, unable to load page.")
                raise e  # Raise the exception after max retries



def safe_get(driver, url, retries=3, delay=15):
    for attempt in range(retries):
        try:
            driver.get(url)
            return  # Exit if successful
        except Exception as e:
            print(f"Attempt {attempt + 1} failed: {e}")
        
        time.sleep(delay)  # Wait before retrying
    print("Failed to load the page after multiple attempts.")


def download_all_images(image_urls, dictionary, data_id):
    product_root_folder = f"{dictionary['search']}"
    for index, image_url in enumerate(image_urls):
        # image_url = image.get_attribute("src")
        print(f"Image {index + 1}: {image_url}")
        folder_path = os.path.join(product_root_folder, dictionary["search"] + " images",str(data_id))
        if not os.path.exists(folder_path):
            # print("path doesn't exists, i ll create")

            # folder_path = os.path.join({product_root_folder},{data_id})
            os.makedirs(folder_path)
        # print("path exists i wont created it")
        download_image(image_url, os.path.join(folder_path,str(index)))
    return folder_path


# def replace_names_with_ids(seller_csv, items_csv):

#     seller_csv_path = "Sellers.csv"
#     seller_df = pd.read_csv(seller_csv)
#     items_df = pd.read_csv(items_csv)

#     names_to_convert = list(items_df[items_df['SellerId'].apply(lambda x: not str(x).replace('.', '', 1).isdigit())])

#     seller_names_saved = seller_df.loc["SellerName"].values

#     max_id = seller_df["SellerId"].max()

#     for name in names_to_convert:
#         if name in seller_names_saved:
#             seller_id = seller_df.loc[seller_df["SellerName" == name, "SellerId"]]
#             items_df[items_df["SellerId"] == name] = seller_id
#         else:
#             max_id += 1




#     #checking if the seller is already saved or not and adding it if not
#     if seller_name in seller_df.loc["SellerName"].values:
#         seller_id = seller_df.loc[seller_df["SellerName" == seller_name, "SellerId"]]
#     else:
#         max_id = seller_df["SellerId"].max()
#         seller_id = max_id + 1
#         new_row = {
#             "SellerId": seller_id,
#             "SellerName": seller_name,
#             "Location": location,
#             "ReviewsCount": reviews_count,
#             "Stars": stars
#         }
#         seller_df.append(new_row, ignore_index=True) 